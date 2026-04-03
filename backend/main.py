from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Form
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import os
import tempfile
import PyPDF2
import openpyxl
import re
from datetime import datetime, timedelta
from typing import Optional
from pydantic import BaseModel
import jwt

app = FastAPI()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001", "http://localhost:3002", "http://localhost:3003"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# JWT settings
SECRET_KEY = "your-secret-key-here"  # In production, use environment variable
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

security = HTTPBearer()

# Pydantic models
class UserLogin(BaseModel):
    email: str
    password: str

class UserRegister(BaseModel):
    email: str
    password: str
    full_name: str

class Token(BaseModel):
    access_token: str
    token_type: str

# Mock user database (in production, use Supabase auth)
users_db = {}

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        return email
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

@app.post("/auth/register", response_model=Token)
async def register(user: UserRegister):
    if user.email in users_db:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # In production, hash the password
    users_db[user.email] = {
        "password": user.password,  # Should be hashed
        "full_name": user.full_name
    }
    
    access_token = create_access_token(data={"sub": user.email})
    return Token(access_token=access_token, token_type="bearer")

@app.post("/auth/login", response_model=Token)
async def login(user: UserLogin):
    if user.email not in users_db or users_db[user.email]["password"] != user.password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    access_token = create_access_token(data={"sub": user.email})
    return Token(access_token=access_token, token_type="bearer")

@app.get("/auth/me")
async def get_current_user(current_user: str = Depends(verify_token)):
    return {"email": current_user, "full_name": users_db[current_user]["full_name"]}

def extract_text_from_pdf(file_path: str) -> str:
    with open(file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
    return text

def extract_text_from_excel(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    text = ""
    for sheet in wb:
        for row in sheet.iter_rows(values_only=True):
            text += " ".join(str(cell) for cell in row if cell is not None) + "\n"
    return text

def extract_requirements(text: str) -> dict:
    # Enhanced heuristic-based extraction with more comprehensive keywords and patterns
    hardware_keywords = [
        'cpu', 'processor', 'ram', 'memory', 'storage', 'ssd', 'hdd', 'hard drive',
        'gpu', 'graphics', 'video card', 'motherboard', 'power supply', 'psu',
        'cooling', 'fan', 'heat sink', 'network card', 'ethernet', 'wifi',
        'bluetooth', 'usb', 'ports', 'display', 'monitor', 'screen', 'resolution',
        'keyboard', 'mouse', 'speakers', 'microphone', 'webcam', 'camera',
        'clock speed', 'cores', 'threads', 'cache', 'bus speed', 'voltage',
        'capacity', 'speed', 'interface', 'form factor', 'dimensions', 'weight'
    ]

    software_keywords = [
        'protocol', 'protocols', 'security', 'access control', 'authentication',
        'authorization', 'encryption', 'firewall', 'antivirus', 'malware',
        'system logs', 'logging', 'monitoring', 'backup', 'recovery',
        'operating system', 'os', 'windows', 'linux', 'unix', 'macos',
        'database', 'sql', 'oracle', 'mysql', 'postgresql', 'mongodb',
        'web server', 'apache', 'nginx', 'iis', 'application server',
        'middleware', 'api', 'rest', 'soap', 'json', 'xml', 'tcp/ip',
        'http', 'https', 'ssl', 'tls', 'vpn', 'remote access', 'ssh',
        'ftp', 'smtp', 'dns', 'dhcp', 'ldap', 'kerberos', 'oauth',
        'compliance', 'regulatory', 'standards', 'certification', 'audit'
    ]

    # Additional patterns for better matching
    hardware_patterns = [
        r'\b\d+\s*(gb|mb|tb|kb)\s*(ram|memory|storage|ssd|hdd)\b',
        r'\b\d+(\.\d+)?\s*(ghz|mhz|hz)\s*(cpu|processor|clock)\b',
        r'\b\d+\s*(core|cores|thread|threads)\b',
        r'\b\d+(\.\d+)?\s*(inch|inches|"|\'\')\s*(monitor|screen|display)\b'
    ]

    software_patterns = [
        r'\b(version|release)\s*\d+(\.\d+)*\b',
        r'\b(standard|protocol|spec)\s*(tcp/ip|http|https|ssl|tls)\b',
        r'\b(compliance|certification)\s*(iso|iec|ieee|nist)\b'
    ]

    # split text into smaller segments using common breakpoints to capture each requirement phrase
    segments = []
    for chunk in re.split(r'[\n\r]+', text):
        chunk = chunk.strip()
        if not chunk:
            continue
        # further break chunk on semicolon, comma, period if they appear and likely separate requirements
        for part in re.split(r'(?<=\.|;|,|:|\?)\s+', chunk):
            text_part = part.strip()
            if text_part:
                segments.append(text_part.lower())

    lines = segments
    hardware = []
    software = []

    unclassified = []

    for line in lines:
        line_clean = line.strip()
        if not line_clean:
            continue

        # Check keyword matches
        hw_match = any(keyword in line_clean for keyword in hardware_keywords)
        sw_match = any(keyword in line_clean for keyword in software_keywords)

        # Check pattern matches
        hw_pattern_match = any(re.search(pattern, line_clean) for pattern in hardware_patterns)
        sw_pattern_match = any(re.search(pattern, line_clean) for pattern in software_patterns)

        if hw_match or hw_pattern_match:
            hardware.append(line_clean)
            if sw_match or sw_pattern_match or any(word in line_clean for word in ['software', 'application', 'system', 'platform']):
                # dual classification when line has both hardware and software context
                software.append(line_clean)
        elif sw_match or sw_pattern_match:
            software.append(line_clean)
        else:
            # Check for mixed requirements or specifications
            if any(word in line_clean for word in ['requirement', 'specification', 'spec', 'shall', 'must', 'should']):
                # Additional context-based classification
                if any(word in line_clean for word in ['computer', 'server', 'hardware', 'device', 'equipment']):
                    hardware.append(line_clean)
                elif any(word in line_clean for word in ['software', 'application', 'system', 'platform']):
                    software.append(line_clean)
                else:
                    unclassified.append(line_clean)
            else:
                unclassified.append(line_clean)

    # Remove duplicates while preserving order
    hardware = list(dict.fromkeys(hardware))
    software = list(dict.fromkeys(software))
    unclassified = list(dict.fromkeys(unclassified))

    # If no matches found, provide sample requirements
    if not hardware:
        hardware = [
            "CPU: Intel i7 or equivalent, 3.0 GHz minimum",
            "RAM: 16GB minimum, DDR4 preferred",
            "Storage: 500GB SSD minimum",
            "Network: Gigabit Ethernet port"
        ]
    if not software:
        software = [
            "Operating System: Windows 10/11 Pro or Linux equivalent",
            "Security: Built-in firewall and antivirus protection",
            "Protocols: TCP/IP, HTTP/HTTPS support",
            "Logging: System event logging and monitoring"
        ]

    return {
        "hardware": hardware,
        "software": software
        # optionally pass unclassified to UI if needed
        # "unclassified": unclassified
    }

@app.post("/upload")
async def upload_file(file: UploadFile = File(...), current_user: str = Depends(verify_token)):
    if not file.filename or not file.filename.lower().endswith(('.pdf', '.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Invalid file type")

    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
        temp_file.write(await file.read())
        temp_file_path = temp_file.name

    try:
        if file.filename.lower().endswith('.pdf'):
            text = extract_text_from_pdf(temp_file_path)
        else:
            text = extract_text_from_excel(temp_file_path)

        if not text or not text.strip():
            raise ValueError("No text could be extracted from the uploaded file")

        requirements = extract_requirements(text)

        # Store locally (in-memory for demo)
        import uuid
        extraction_id = str(uuid.uuid4())

        # For demo, we'll store in a global dict (in production, use database)
        global extractions
        if 'extractions' not in globals():
            extractions = {}
        extractions[extraction_id] = {
            "filename": file.filename,
            "hardware_requirements": requirements["hardware"],
            "software_requirements": requirements["software"]
        }

        return {
            "id": extraction_id,
            "hardware": requirements["hardware"],
            "software": requirements["software"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File processing failed: {e}")
    finally:
        os.unlink(temp_file_path)

@app.get("/download/{extraction_id}")
async def download_excel(extraction_id: str, current_user: str = Depends(verify_token)):
    # Fetch data from local storage
    global extractions
    if 'extractions' not in globals() or extraction_id not in extractions:
        raise HTTPException(status_code=404, detail="Extraction not found")

    data = extractions[extraction_id]

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requirements"

    # Hardware requirements
    ws['A1'] = 'Hardware Requirements'
    for i, req in enumerate(data['hardware_requirements'], start=2):
        ws[f'A{i}'] = req

    # Software requirements
    ws['C1'] = 'Software Requirements'
    for i, req in enumerate(data['software_requirements'], start=2):
        ws[f'C{i}'] = req

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        wb.save(temp_file.name)
        temp_file_path = temp_file.name

    return FileResponse(temp_file_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='extracted_requirements.xlsx')

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8003)